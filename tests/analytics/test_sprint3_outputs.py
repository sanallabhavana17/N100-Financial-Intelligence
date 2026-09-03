from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook


# ============================================================
# PATHS
# ============================================================

BASE_DIR = Path(__file__).resolve().parents[2]

PEER_CSV = BASE_DIR / "output" / "peer_percentile_table.csv"
PEER_XLSX = BASE_DIR / "output" / "peer_comparison.xlsx"


# ============================================================
# EXPECTED VALUES
# ============================================================

EXPECTED_PEER_GROUPS = 11
EXPECTED_ROWS = 659

EXPECTED_PEER_GROUP_NAMES = {
    "Automobiles",
    "Consumer Finance",
    "FMCG",
    "IT Services",
    "Life Insurance",
    "Oil & Gas",
    "Pharmaceuticals",
    "Power & Utilities",
    "Private Banks",
    "Public Sector Banks",
    "Steel",
}


# ============================================================
# FIXTURE
# ============================================================

@pytest.fixture(scope="module")
def peer_data():
    assert PEER_CSV.exists(), (
        f"Peer percentile CSV not found: {PEER_CSV}"
    )

    return pd.read_csv(PEER_CSV)


# ============================================================
# CSV DATA QUALITY TESTS
# ============================================================

def test_peer_csv_exists():
    assert PEER_CSV.exists()


def test_peer_csv_row_count(peer_data):
    assert len(peer_data) == EXPECTED_ROWS


def test_peer_csv_peer_group_count(peer_data):
    peer_groups = set(
        peer_data["peer_group_name"]
        .dropna()
        .astype(str)
        .unique()
    )

    assert len(peer_groups) == EXPECTED_PEER_GROUPS
    assert peer_groups == EXPECTED_PEER_GROUP_NAMES


def test_peer_csv_required_columns(peer_data):
    required_columns = {
        "peer_group_name",
        "company_id",
        "year",
        "is_benchmark",
        "peer_rank",
        "peer_group_size",
        "peer_composite_percentile",
        "benchmark_company_id",
        "benchmark_composite_percentile",
        "vs_benchmark_percentile",
        "above_benchmark",
        "composite_quality_score",
        "composite_quality_score_percentile",
    }

    missing_columns = required_columns - set(peer_data.columns)

    assert not missing_columns, (
        f"Missing required columns: {sorted(missing_columns)}"
    )


def test_peer_percentiles_are_valid(peer_data):
    """
    Validate genuine percentile columns.

    vs_benchmark_percentile is intentionally excluded because
    it represents the difference versus the benchmark and can
    legitimately contain negative values.
    """

    percentile_columns = [
        column
        for column in peer_data.columns
        if column.endswith("_percentile")
        and column != "vs_benchmark_percentile"
    ]

    assert percentile_columns

    for column in percentile_columns:
        values = pd.to_numeric(
            peer_data[column],
            errors="coerce",
        ).dropna()

        assert ((values >= 0) & (values <= 100)).all(), (
            f"Invalid percentile values found in {column}"
        )


def test_peer_group_company_year_has_no_duplicates(peer_data):
    key_columns = [
        "peer_group_name",
        "company_id",
        "year",
    ]

    duplicates = peer_data.duplicated(
        subset=key_columns,
        keep=False,
    )

    assert not duplicates.any(), (
        "Duplicate peer/company/year records found"
    )


def test_peer_rank_values_are_valid(peer_data):
    ranks = pd.to_numeric(
        peer_data["peer_rank"],
        errors="coerce",
    )

    group_sizes = pd.to_numeric(
        peer_data["peer_group_size"],
        errors="coerce",
    )

    valid_rows = ranks.notna() & group_sizes.notna()

    assert (ranks[valid_rows] >= 1).all()
    assert (ranks[valid_rows] <= group_sizes[valid_rows]).all()


# ============================================================
# EXCEL OUTPUT TESTS
# ============================================================

def test_peer_excel_exists():
    assert PEER_XLSX.exists()


def test_peer_excel_has_11_sheets():
    assert PEER_XLSX.exists()

    workbook = load_workbook(
        PEER_XLSX,
        read_only=True,
    )

    try:
        assert len(workbook.sheetnames) == EXPECTED_PEER_GROUPS

        assert set(workbook.sheetnames) == EXPECTED_PEER_GROUP_NAMES
    finally:
        workbook.close()


def test_peer_excel_has_659_data_rows():
    assert PEER_XLSX.exists()

    workbook = load_workbook(
        PEER_XLSX,
        read_only=True,
    )

    try:
        total_rows = 0

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            # Row 4 contains the header.
            data_rows = max(
                worksheet.max_row - 4,
                0,
            )

            total_rows += data_rows

        assert total_rows == EXPECTED_ROWS

    finally:
        workbook.close()