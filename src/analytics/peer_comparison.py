"""
NIFTY100 FINANCIAL INTELLIGENCE
SPRINT 3 - DAY 20
PEER COMPARISON EXCEL EXPORT

Purpose:
    Create a formatted Excel workbook containing one sheet
    for each peer group.

Input:
    output/peer_percentile_table.csv

Output:
    output/peer_comparison.xlsx

Requirements:
    - 11 peer-group sheets
    - Company-level peer comparison
    - Peer rank
    - Peer percentile
    - Benchmark comparison
    - Relevant financial KPIs
    - Formatted and sortable workbook
"""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter


# ============================================================
# PATHS
# ============================================================

BASE_DIR = Path(__file__).resolve().parents[2]

INPUT_FILE = (
    BASE_DIR
    / "output"
    / "peer_percentile_table.csv"
)

OUTPUT_FILE = (
    BASE_DIR
    / "output"
    / "peer_comparison.xlsx"
)


# ============================================================
# CONFIGURATION
# ============================================================

EXPECTED_PEER_GROUPS = 11

MAX_EXCEL_SHEET_NAME_LENGTH = 31


# Columns to display in each peer-group sheet.
# The script automatically keeps only columns that exist.
DISPLAY_COLUMNS = [
    "peer_group_name",
    "company_id",
    "year",
    "is_benchmark",
    "peer_rank",
    "peer_group_size",
    "peer_composite_percentile",
    "benchmark_company_id",
    "benchmark_composite_percentile",
    "vs_benchmark_percentile",
    "above_benchmark",

    # Financial metrics
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "return_on_assets_pct",
    "net_profit_margin_pct",
    "operating_profit_margin_pct",
    "debt_to_equity",
    "interest_coverage",
    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "eps_cagr_5yr",
    "free_cash_flow_cr",
    "asset_turnover",
    "composite_quality_score",

    # Percentile metrics
    "return_on_equity_pct_percentile",
    "return_on_capital_employed_pct_percentile",
    "return_on_assets_pct_percentile",
    "net_profit_margin_pct_percentile",
    "operating_profit_margin_pct_percentile",
    "debt_to_equity_percentile",
    "interest_coverage_percentile",
    "revenue_cagr_5yr_percentile",
    "pat_cagr_5yr_percentile",
    "eps_cagr_5yr_percentile",
    "free_cash_flow_cr_percentile",
    "asset_turnover_percentile",
    "composite_quality_score_percentile",
]


# ============================================================
# STYLING
# ============================================================

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78"
)

HEADER_FONT = Font(
    bold=True,
    color="FFFFFF"
)

TITLE_FONT = Font(
    bold=True,
    size=14
)

SUBTITLE_FONT = Font(
    italic=True,
    size=10
)

BENCHMARK_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFF2CC"
)

THIN_BORDER = Border(
    bottom=Side(
        style="thin",
        color="D9E1F2"
    )
)


# ============================================================
# LOAD DATA
# ============================================================

def load_peer_data():
    """
    Load the D18 peer percentile table.
    """

    print("Loading peer percentile data...")
    print(f"Input: {INPUT_FILE}")

    if not INPUT_FILE.exists():
        raise FileNotFoundError(
            f"Input file not found: {INPUT_FILE}"
        )

    df = pd.read_csv(INPUT_FILE)

    print(f"Rows loaded: {len(df):,}")

    if "peer_group_name" not in df.columns:
        raise ValueError(
            "peer_group_name column is missing."
        )

    if "company_id" not in df.columns:
        raise ValueError(
            "company_id column is missing."
        )

    peer_groups = (
        df["peer_group_name"]
        .dropna()
        .astype(str)
        .unique()
    )

    print(
        f"Peer groups found: "
        f"{len(peer_groups)}"
    )

    print(
        f"Companies represented: "
        f"{df['company_id'].nunique()}"
    )

    return df


# ============================================================
# PREPARE DATA
# ============================================================

def prepare_data(df):
    """
    Prepare and clean data before exporting.
    """

    data = df.copy()

    if "year" in data.columns:
        data["year"] = pd.to_numeric(
            data["year"],
            errors="coerce"
        )

    # Convert numeric-looking columns.
    for column in data.columns:
        if column in {
            "peer_group_name",
            "company_id",
            "benchmark_company_id",
            "above_benchmark",
        }:
            continue

        data[column] = pd.to_numeric(
            data[column],
            errors="ignore"
        )

    return data


# ============================================================
# SHEET NAME
# ============================================================

def make_sheet_name(peer_group, used_names):
    """
    Create a valid and unique Excel sheet name.
    """

    name = str(peer_group)

    # Excel-invalid characters.
    for char in [
        "\\",
        "/",
        "*",
        "?",
        ":",
        "[",
        "]",
    ]:
        name = name.replace(char, "_")

    name = name.strip()

    if not name:
        name = "Peer_Group"

    name = name[
        :MAX_EXCEL_SHEET_NAME_LENGTH
    ]

    original = name
    counter = 1

    while name in used_names:
        suffix = f"_{counter}"

        name = (
            original[
                :MAX_EXCEL_SHEET_NAME_LENGTH
                - len(suffix)
            ]
            + suffix
        )

        counter += 1

    used_names.add(name)

    return name


# ============================================================
# WRITE SHEET
# ============================================================

def write_peer_sheet(
    writer,
    peer_group,
    peer_df,
    sheet_name,
):
    """
    Write one peer group to one Excel sheet.
    """

    available_columns = [
        column
        for column in DISPLAY_COLUMNS
        if column in peer_df.columns
    ]

    sheet_df = peer_df[
        available_columns
    ].copy()

    # Sort benchmark first, then peer rank.
    sort_columns = []

    if "is_benchmark" in sheet_df.columns:
        sort_columns.append(
            "is_benchmark"
        )

    if "peer_rank" in sheet_df.columns:
        sort_columns.append(
            "peer_rank"
        )

    if sort_columns:
        ascending = [
            False
            if column == "is_benchmark"
            else True
            for column in sort_columns
        ]

        sheet_df = sheet_df.sort_values(
            sort_columns,
            ascending=ascending
        )

    # Write data starting at row 4.
    sheet_df.to_excel(
        writer,
        sheet_name=sheet_name,
        index=False,
        startrow=3
    )


# ============================================================
# FORMAT WORKBOOK
# ============================================================

def format_workbook(
    output_file,
    peer_groups,
):
    """
    Apply professional formatting to the workbook.
    """

    print()
    print("Formatting Excel workbook...")

    wb = load_workbook(output_file)

    for ws in wb.worksheets:

        peer_group = ws.title

        # ----------------------------------------------------
        # Title
        # ----------------------------------------------------

        ws["A1"] = (
            f"NIFTY100 Peer Comparison - "
            f"{peer_group}"
        )

        ws["A1"].font = TITLE_FONT

        ws["A2"] = (
            "Peer-ranked financial KPI comparison "
            "with benchmark reference"
        )

        ws["A2"].font = SUBTITLE_FONT

        # ----------------------------------------------------
        # Header row
        # ----------------------------------------------------

        header_row = 4

        for cell in ws[header_row]:

            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )
            cell.border = THIN_BORDER

        ws.row_dimensions[
            header_row
        ].height = 36

        # ----------------------------------------------------
        # Freeze panes
        # ----------------------------------------------------

        ws.freeze_panes = "C5"

        # ----------------------------------------------------
        # Auto filter
        # ----------------------------------------------------

        if ws.max_row >= header_row:
            ws.auto_filter.ref = (
                f"A{header_row}:"
                f"{get_column_letter(ws.max_column)}"
                f"{ws.max_row}"
            )

        # ----------------------------------------------------
        # Column widths
        # ----------------------------------------------------

        for column_cells in ws.columns:

            column_letter = (
                get_column_letter(
                    column_cells[0].column
                )
            )

            max_length = 0

            for cell in column_cells:

                try:
                    value_length = len(
                        str(cell.value)
                    )
                except Exception:
                    value_length = 0

                max_length = max(
                    max_length,
                    value_length
                )

            width = min(
                max(max_length + 2, 12),
                28
            )

            ws.column_dimensions[
                column_letter
            ].width = width

        # ----------------------------------------------------
        # Format numeric cells
        # ----------------------------------------------------

        for row in ws.iter_rows(
            min_row=header_row + 1,
            max_row=ws.max_row
        ):

            for cell in row:

                if isinstance(
                    cell.value,
                    (int, float)
                ):

                    cell.number_format = (
                        "0.00"
                    )

        # ----------------------------------------------------
        # Highlight benchmark rows
        # ----------------------------------------------------

        headers = {
            cell.value: cell.column
            for cell in ws[header_row]
        }

        benchmark_column = headers.get(
            "is_benchmark"
        )

        if benchmark_column:

            for row_number in range(
                header_row + 1,
                ws.max_row + 1
            ):

                cell = ws.cell(
                    row=row_number,
                    column=benchmark_column
                )

                value = str(
                    cell.value
                ).lower()

                if value in {
                    "true",
                    "1",
                    "yes",
                }:

                    for column_number in range(
                        1,
                        ws.max_column + 1
                    ):

                        ws.cell(
                            row=row_number,
                            column=column_number
                        ).fill = BENCHMARK_FILL

        # ----------------------------------------------------
        # Conditional formatting for percentile columns
        # ----------------------------------------------------

        percentile_columns = []

        for cell in ws[header_row]:

            if (
                cell.value
                and "percentile"
                in str(cell.value).lower()
            ):
                percentile_columns.append(
                    cell.column
                )

        for column_number in percentile_columns:

            column_letter = (
                get_column_letter(
                    column_number
                )
            )

            cell_range = (
                f"{column_letter}"
                f"{header_row + 1}:"
                f"{column_letter}"
                f"{ws.max_row}"
            )

            ws.conditional_formatting.add(
                cell_range,
                ColorScaleRule(
                    start_type="min",
                    start_color="F8696B",
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FFEB84",
                    end_type="max",
                    end_color="63BE7B"
                )
            )

        # ----------------------------------------------------
        # Row heights
        # ----------------------------------------------------

        for row_number in range(
            header_row + 1,
            ws.max_row + 1
        ):
            ws.row_dimensions[
                row_number
            ].height = 20

    # --------------------------------------------------------
    # Save
    # --------------------------------------------------------

    wb.save(output_file)

    print(
        f"Formatted workbook saved: "
        f"{output_file}"
    )


# ============================================================
# VALIDATE WORKBOOK
# ============================================================

def validate_workbook(
    output_file,
    expected_peer_groups,
):
    """
    Validate the generated workbook.
    """

    print()
    print("OUTPUT VALIDATION")
    print("-" * 60)

    if not output_file.exists():
        raise FileNotFoundError(
            "Excel output file was not created."
        )

    wb = load_workbook(
        output_file,
        read_only=True
    )

    sheet_names = wb.sheetnames

    print(
        f"Sheets generated: "
        f"{len(sheet_names)}"
    )

    print(
        "Sheet names:"
    )

    for sheet in sheet_names:
        print(f"  - {sheet}")

    # --------------------------------------------------------
    # Sheet count
    # --------------------------------------------------------

    if len(sheet_names) != expected_peer_groups:
        raise ValueError(
            f"Expected {expected_peer_groups} sheets, "
            f"but found {len(sheet_names)}."
        )

    # --------------------------------------------------------
    # Validate each sheet has data
    # --------------------------------------------------------

    total_data_rows = 0

    for sheet_name in sheet_names:

        ws = wb[sheet_name]

        # Header at row 4.
        if ws.max_row < 5:
            raise ValueError(
                f"Sheet '{sheet_name}' "
                "contains no company data."
            )

        data_rows = ws.max_row - 4

        total_data_rows += data_rows

        print(
            f"  {sheet_name}: "
            f"{data_rows} company rows"
        )

    print(
        f"Total company rows exported: "
        f"{total_data_rows}"
    )

    wb.close()

    print()
    print("Workbook validation: PASSED")


# ============================================================
# MAIN
# ============================================================

def main():

    print("=" * 60)
    print("NIFTY100 PEER COMPARISON EXCEL EXPORT")
    print("SPRINT 3 - DAY 20")
    print("=" * 60)

    # --------------------------------------------------------
    # 1. Load data
    # --------------------------------------------------------

    df = load_peer_data()

    # --------------------------------------------------------
    # 2. Prepare data
    # --------------------------------------------------------

    df = prepare_data(df)

    # --------------------------------------------------------
    # 3. Get peer groups
    # --------------------------------------------------------

    peer_groups = sorted(
        df["peer_group_name"]
        .dropna()
        .astype(str)
        .unique()
    )

    if len(peer_groups) != EXPECTED_PEER_GROUPS:
        raise ValueError(
            f"Expected {EXPECTED_PEER_GROUPS} peer groups, "
            f"but found {len(peer_groups)}."
        )

    print()
    print("Peer groups:")
    for peer_group in peer_groups:
        count = (
            df.loc[
                df["peer_group_name"].astype(str)
                == peer_group,
                "company_id"
            ]
            .nunique()
        )

        print(
            f"  - {peer_group}: "
            f"{count} companies"
        )

    # --------------------------------------------------------
    # 4. Create Excel workbook
    # --------------------------------------------------------

    OUTPUT_FILE.parent.mkdir(
        parents=True,
        exist_ok=True
    )

    used_sheet_names = set()

    print()
    print("Creating Excel workbook...")

    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl"
    ) as writer:

        for peer_group in peer_groups:

            peer_df = df[
                df["peer_group_name"]
                .astype(str)
                == peer_group
            ].copy()

            sheet_name = make_sheet_name(
                peer_group,
                used_sheet_names
            )

            write_peer_sheet(
                writer,
                peer_group,
                peer_df,
                sheet_name
            )

            print(
                f"  Created sheet: "
                f"{sheet_name}"
            )

    # --------------------------------------------------------
    # 5. Format workbook
    # --------------------------------------------------------

    format_workbook(
        OUTPUT_FILE,
        peer_groups
    )

    # --------------------------------------------------------
    # 6. Validate workbook
    # --------------------------------------------------------

    validate_workbook(
        OUTPUT_FILE,
        EXPECTED_PEER_GROUPS
    )

    # --------------------------------------------------------
    # 7. Final message
    # --------------------------------------------------------

    print()
    print("=" * 60)
    print("D20 PEER COMPARISON EXPORT COMPLETE")
    print("=" * 60)

    print(
        f"Output: {OUTPUT_FILE}"
    )

    print(
        f"Peer groups: {len(peer_groups)}"
    )

    print(
        f"Sheets: {len(peer_groups)}"
    )

    print("=" * 60)


if __name__ == "__main__":
    main()